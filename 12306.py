# -*- coding: utf-8 -*-
"""
import openpyxl
import re

def read_excel():
    # 打开一张表
    wb=openpyxl.load_workbook('D:\\12306.xlsx')
    
    ws = wb.create_sheet(wb, '12306')


    #xl_sheet=wb.active # 获取活跃表单
    rows=ws.max_row
    i=1
    str1="@"
    str2="$"
    while i<= rows:
        #将口令长度限制在6~25之间
        strnow=ws.cell(i,7).value
        if type(strnow) == int :
            strnow=str(strnow)

        if len(strnow)<6 or len(strnow)>25 :
            ws.delete_rows(i)
            continue 
        
        #将@换成a，将$换成s
        if str1 in ws.cell(i,6).value :
            ws.cell(i,6).value.replace('@','a')

        if str2 in ws.cell(i,6).value :
            ws.cell(i,6).value.replace('$','s')
    
        #将带有其他字符的口令里面的特殊字符删掉？---不处理特殊字符
        i+=1

    wb.save('12306')
    print(i)


if __name__ == '__main__':
    read_excel()
"""
import pandas as pd
X=pd.read_csv("D:\\12306.csv" ,columns=['0','1','2','3','4','5','6'])
X=X.values;
t=[]
for i in range(X.shape[0]):
    if len(X[i][6])>6 and len(X[i][6])<25:
        X[i][6].replace('$','s')
        X[i][6].replace('@','a')
        t.append(X[i]);
T=pd.DataFrame(t);
T.to_csv('D://12306cleaned.csv');
